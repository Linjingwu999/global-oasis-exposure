from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
import yaml

from .estimands import point_result, q_pass, sample_for_estimand
from .io_utils import atomic_bytes, atomic_csv, atomic_json, sha256
from .sensitivity import futurepop_accounting, validate_locked_sensitivity
from .spatial_blocks import validate_block_mapping


REQUIRED_FILES = {
    "analysis_config": "config/analysis.yml",
    "estimand_contract": "config/estimands.csv",
    "minimum_input": "data/analysis_input_minimal.csv",
    "primary_estimands": "data/primary_estimands_21.csv",
    "numeric_facts": "data/master_numeric_facts_29.csv",
    "source_data": "data/Source_Data_CEE_v1.xlsx",
    "data_dictionary": "data/data_dictionary.csv",
    "source_manifest": "data/source_product_manifest.csv",
}

AUTHORITATIVE_RELEASE_SHA256 = {
    "analysis_config": "29FB716EB52F1504FA66F8921C6B083192FD931666E8E5302DE3CB6332D6CA52",
    "estimand_contract": "7B6438B67158D16AA1CDA60B80B9CF52BD6611E45255FDAA56C26DBA78D38B89",
    "minimum_input": "F80B6FA4F02BF1DF9E487C9E08564CA5346502C5EC2E64D7522210CE54D8F10F",
    "primary_estimands": "675578AC85087C88A5EA399719884B8B735DA1D168D997BBC666C15592F91C66",
    "numeric_facts": "8E9A903531E9FA5FA384E8F8E25C205276882D347D1213EF6E98E7AA535AF816",
    "source_data": "9C40550F6ADA12870A6904CEEA975489998347BAB8F7F756C2074D5F55F59B1F",
    "data_dictionary": "BDE34F7121BF4B9C896D4976CB2ECA594AA91B26F1D86F426CA9E2AE68B70D43",
    "source_manifest": "6E5EC006D8372CE418BBABC251D569DCA5F2249D18C90B6832941C27E2848511",
}

EXPECTED_CLASS_COUNTS = {
    "BWh oases": 1822,
    "BWk oases": 742,
    "BWh/BWk oases": 131,
    "non-BW oases": 748,
}
EXPECTED_CONTINENT_COUNTS = {
    "Asia": 1593,
    "Africa": 1096,
    "North America": 591,
    "South America": 123,
    "Oceania": 40,
}
EXPECTED_STRICT_BLOCK_COUNTS = {"250": 352, "500": 148, "1000": 65}
EXPECTED_DOMAIN_FAMILY_SIZES = {"social": 2, "water": 3, "utci": 4, "nex": 12}
EXPECTED_FACT_IDS = [
    *[f"F04-PRIMARY-{index:02d}" for index in range(1, 22)],
    "F04-CONTEXT-01",
    "F04-CONTEXT-02",
    "F04-CONTEXT-03",
    "F04-FUTUREPOP-01",
    "F04-FUTUREPOP-02",
    "F04-FUTUREPOP-03",
    "F04-FUTUREPOP-04",
    "F04-FUTUREPOP-05",
]


class ValidationFailure(RuntimeError):
    pass


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise ValidationFailure(message)


def _close(actual: object, expected: object) -> bool:
    return bool(
        np.isclose(
            float(actual),
            float(expected),
            rtol=1e-10,
            atol=1e-10,
            equal_nan=True,
        )
    )


def _workbook_qa(
    path: Path, primary: pd.DataFrame, facts: pd.DataFrame
) -> dict[str, object]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    required_sheets = {
        "README",
        "Data_Dictionary",
        "Table2_Estimands",
        "NumericFacts29",
        "SuppTableS3_Estimands",
        "SuppTableS7_Products",
    }
    _require(required_sheets.issubset(workbook.sheetnames), "Source Data sheets are missing")
    def worksheet_frame(name: str) -> pd.DataFrame:
        rows = list(workbook[name].values)
        _require(bool(rows), f"Source Data sheet is empty: {name}")
        return pd.DataFrame(rows[1:], columns=rows[0])

    table2 = worksheet_frame("Table2_Estimands")
    numeric29 = worksheet_frame("NumericFacts29")
    try:
        pd.testing.assert_frame_equal(
            table2.reset_index(drop=True),
            primary.reset_index(drop=True),
            check_dtype=False,
            rtol=1e-12,
            atol=1e-12,
        )
        pd.testing.assert_frame_equal(
            numeric29.reset_index(drop=True),
            facts.reset_index(drop=True),
            check_dtype=False,
            rtol=1e-12,
            atol=1e-12,
        )
    except AssertionError as exc:
        raise ValidationFailure(f"Source Data table mismatch: {exc}") from exc
    ai_scaling_evidence = str(workbook["SuppTableS7_Products"]["E4"].value)
    _require("0.0001" in ai_scaling_evidence, "Source Data AI scale-factor evidence is missing")

    formula_count = 0
    privacy_hits: list[dict[str, str]] = []
    patterns = {
        "windows_absolute_path": re.compile(r"(?<![A-Za-z0-9])[A-Za-z]:\\"),
        "posix_home_path": re.compile(r"/(?:Users|home)/[^/\s]+/"),
        "sandbox_path": re.compile(r"sandbox" + r":/|/mnt" + r"/data/"),
        "email": re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.I),
    }
    cell_count = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_count += 1
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    formula_count += 1
                if isinstance(cell.value, str):
                    for kind, pattern in patterns.items():
                        if pattern.search(cell.value):
                            privacy_hits.append(
                                {"sheet": worksheet.title, "cell": cell.coordinate, "kind": kind}
                            )
    workbook.close()
    _require(formula_count == 0, f"Source Data contains {formula_count} formulas")
    _require(not privacy_hits, f"Source Data privacy hits: {privacy_hits[:5]}")
    return {
        "sheet_count": len(workbook.sheetnames),
        "required_sheets_present": True,
        "nonempty_cells_scanned": cell_count,
        "formula_count": formula_count,
        "privacy_hits": privacy_hits,
        "table2_matches_primary_estimands": True,
        "numericfacts29_matches_csv": True,
        "ai_scale_factor": 0.0001,
        "ai_scaling_evidence": ai_scaling_evidence,
    }


def _append_log(path: Path, message: str) -> None:
    timestamp = datetime.now(timezone.utc).isoformat(timespec="seconds")
    existing = path.read_text(encoding="utf-8") if path.exists() else ""
    atomic_bytes(path, f"{existing}{timestamp} {message}\n".encode("utf-8"))


def _load_interrupted_state(
    path: Path, fingerprint: dict[str, str], resume: bool
) -> dict[str, Any] | None:
    if not resume or not path.exists():
        return None
    try:
        state = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if state.get("resume_fingerprint_sha256") != fingerprint:
        return None
    return state


def _resumable_point_estimands(
    data: pd.DataFrame,
    contract: pd.DataFrame,
    output_dir: Path,
    state: dict[str, Any],
    log_path: Path,
    *,
    resume: bool,
) -> tuple[pd.DataFrame, list[dict[str, str]], int]:
    state_path = output_dir / "reproduction_state.json"
    work = state["work_packages"]["point_estimands"]
    work["status"] = "running"
    state["status"] = "running"
    atomic_json(state_path, state)
    _append_log(log_path, "point_estimands running")
    result_dir = output_dir / "estimand_results"
    results: list[dict[str, object]] = []
    failures: list[dict[str, str]] = []
    resumed_count = 0
    for row in contract.to_dict("records"):
        fact_id = str(row["fact_id"])
        estimand = str(row["estimand"])
        result_path = result_dir / f"{fact_id}.json"
        item = work["items"].setdefault(fact_id, {"status": "pending", "estimand": estimand})
        can_resume = (
            resume
            and item.get("status") == "success"
            and result_path.is_file()
            and item.get("result_sha256") == sha256(result_path)
        )
        if can_resume:
            try:
                stored = json.loads(result_path.read_text(encoding="utf-8"))
                if stored.get("estimand") != estimand:
                    raise ValueError("stored estimand identity changed")
                results.append(stored)
                resumed_count += 1
                _append_log(log_path, f"point_estimand resumed {fact_id}")
                continue
            except (OSError, ValueError, json.JSONDecodeError):
                can_resume = False
        item.update({"status": "running", "estimand": estimand})
        atomic_json(state_path, state)
        _append_log(log_path, f"point_estimand running {fact_id}")
        try:
            sample = sample_for_estimand(data, row)
            result = point_result(sample, row)
            atomic_json(result_path, result)
            item.update(
                {
                    "status": "success",
                    "result_sha256": sha256(result_path),
                    "error": None,
                }
            )
            results.append(result)
            atomic_csv(
                output_dir / "regenerated_point_estimands_partial.csv",
                pd.DataFrame(results),
            )
            _append_log(log_path, f"point_estimand success {fact_id}")
        except Exception as exc:
            item.update({"status": "failed", "error": f"{type(exc).__name__}: {exc}"})
            failures.append({"estimand": estimand, "error": str(exc)})
            _append_log(log_path, f"point_estimand failed {fact_id} {type(exc).__name__}: {exc}")
        atomic_json(state_path, state)
    if failures:
        work["status"] = "failed"
        state["status"] = "failed"
        atomic_json(state_path, state)
        _append_log(log_path, f"point_estimands failed count={len(failures)}")
    else:
        work["status"] = "success"
        atomic_json(state_path, state)
        _append_log(log_path, f"point_estimands success count={len(results)} resumed={resumed_count}")
    return pd.DataFrame(results), failures, resumed_count


def _point_estimate_qa(
    regenerated: pd.DataFrame,
    failures: list[dict[str, str]],
    primary: pd.DataFrame,
    resumed_count: int,
) -> tuple[pd.DataFrame, dict[str, object]]:
    _require(not failures, f"Point-estimation failures: {failures}")
    _require(len(regenerated) == 21, f"Regenerated estimands={len(regenerated)}, expected 21")
    locked = primary.set_index("estimand", drop=False)
    mismatches: list[dict[str, object]] = []
    for row in regenerated.to_dict("records"):
        estimand = str(row["estimand"])
        if estimand not in locked.index:
            mismatches.append({"estimand": estimand, "field": "missing_locked_row"})
            continue
        expected = locked.loc[estimand]
        for field in ("n_bwh", "n_bwk"):
            if int(row[field]) != int(float(expected[field])):
                mismatches.append(
                    {"estimand": estimand, "field": field, "actual": row[field], "expected": expected[field]}
                )
        for field in ("estimate_bwh", "estimate_bwk", "effect_bwh_minus_bwk"):
            if not _close(row[field], expected[field]):
                mismatches.append(
                    {"estimand": estimand, "field": field, "actual": row[field], "expected": expected[field]}
                )
    _require(not mismatches, f"Locked point-estimate mismatches: {mismatches[:5]}")
    return regenerated, {
        "status": "PASS",
        "regenerated": len(regenerated),
        "failures": failures,
        "mismatches": mismatches,
        "resumed_estimators": resumed_count,
        "unrelated_estimators_continue_after_failure": True,
        "bootstrap_recomputed": False,
        "bootstrap_mode": "locked_outputs_verified",
    }


def _locked_output_qa(
    contract: pd.DataFrame, primary: pd.DataFrame, facts: pd.DataFrame
) -> dict[str, object]:
    _require(len(contract) == 21, "Estimand contract does not contain 21 rows")
    _require(len(primary) == 21, "Primary table does not contain 21 rows")
    _require(len(facts) == 29, "Numerical-facts table does not contain 29 rows")
    _require(primary["fact_id"].is_unique, "Primary fact IDs are not unique")
    _require(facts["fact_id"].tolist() == EXPECTED_FACT_IDS, "Numerical fact IDs/order changed")

    contract_columns = [
        "fact_id", "domain", "estimand", "statistic", "unit", "effect_scale", "support_class"
    ]
    contract_projection = contract[contract_columns].astype("string").reset_index(drop=True)
    primary_projection = primary[contract_columns].astype("string").reset_index(drop=True)
    _require(
        contract_projection.equals(primary_projection),
        "Per-estimand contract or support classification differs from the locked primary table",
    )

    support = primary["support_class"].value_counts().to_dict()
    _require(
        support == {"robust": 16, "not_supported": 3, "sensitive": 2},
        f"Support classification changed: {support}",
    )
    families = contract.groupby("domain", sort=False).size().to_dict()
    _require(families == EXPECTED_DOMAIN_FAMILY_SIZES, f"BH family sizes changed: {families}")

    ci_low = pd.to_numeric(primary["ci95_low"], errors="coerce")
    ci_high = pd.to_numeric(primary["ci95_high"], errors="coerce")
    q_values = pd.to_numeric(primary["fdr_q"], errors="coerce")
    _require(bool((ci_low.notna() & ci_high.notna() & (ci_low <= ci_high)).all()), "Invalid confidence intervals")
    _require(bool((q_values.notna() & q_values.between(0, 1)).all()), "Invalid FDR q-values")

    primary_text = pd.read_csv(
        primary.attrs["source_path"], dtype=str, keep_default_na=False
    )
    facts_text = pd.read_csv(facts.attrs["source_path"], dtype=str, keep_default_na=False)
    _require(
        primary_text.equals(facts_text.iloc[:21].reset_index(drop=True)),
        "The first 21 numerical facts are not byte-level table projections of primary outputs",
    )
    sensitivity = validate_locked_sensitivity(primary)
    return {
        "status": "PASS",
        "estimands": len(primary),
        "numeric_facts": len(facts),
        "support_counts": support,
        "bh_family_sizes": families,
        "sensitivity": sensitivity,
        "locked_ci_fdr_verified": True,
        "per_estimand_contract_support_match": True,
        "verification_scope": "authoritative SHA-256 and row-level integrity checks; bootstrap p-values and intervals are not independently recomputed",
    }


def _data_qa(data: pd.DataFrame, contract: pd.DataFrame, facts: pd.DataFrame) -> dict[str, object]:
    _require(data.shape == (3443, 44), f"Minimum input shape changed: {data.shape}")
    _require(data["OasisID"].notna().all() and data["OasisID"].is_unique, "OasisID integrity failed")
    class_counts = data["class_label_en"].value_counts().to_dict()
    continent_counts = data["continent5"].value_counts().to_dict()
    _require(class_counts == EXPECTED_CLASS_COUNTS, f"Class counts changed: {class_counts}")
    _require(continent_counts == EXPECTED_CONTINENT_COUNTS, f"Continent counts changed: {continent_counts}")

    blocks = validate_block_mapping(data)
    _require(
        blocks["occupied_strict_bwh_bwk"] == EXPECTED_STRICT_BLOCK_COUNTS,
        f"Strict-sample block counts changed: {blocks['occupied_strict_bwh_bwk']}",
    )

    utci_pass = int(q_pass(data["utci_qc"]).sum())
    _require(utci_pass == 3434, f"UTCI complete count changed: {utci_pass}")
    nex_methods = data["nex_method"].value_counts().to_dict()
    _require(
        nex_methods == {"polygon_reducer": 2871, "representative_point_fallback": 572},
        f"NEX method counts changed: {nex_methods}",
    )
    futurepop = futurepop_accounting(data)
    _require(
        futurepop == {"coverage_gt0": 3186, "coverage_ge050": 2252, "no_valid": 257},
        f"FuturePop coverage accounting changed: {futurepop}",
    )

    fact_lookup = facts.set_index("fact_id")
    coverage_mask = pd.to_numeric(data["futurepop_coverage"], errors="coerce") > 0
    for index, field in enumerate(
        ("pop_ssp2_2050", "pop_ssp2_2080", "pop_ssp5_2050", "pop_ssp5_2080"),
        start=1,
    ):
        locked = fact_lookup.loc[f"F04-FUTUREPOP-{index:02d}"]
        for label, output_field in (("BWh oases", "estimate_bwh"), ("BWk oases", "estimate_bwk")):
            values = pd.to_numeric(
                data.loc[coverage_mask & data["class_label_en"].eq(label), field],
                errors="coerce",
            )
            _require(
                _close(values.sum(), locked[output_field]),
                f"FuturePop total changed for {field}, {label}",
            )

    ai_values = pd.to_numeric(data["ai_scaled"], errors="coerce")
    _require(int(ai_values.notna().sum()) == 3418, "Scaled AI valid count changed")
    _require(bool(ai_values.dropna().between(0, 10).all()), "Scaled AI range is implausible")

    et0_fields = {
        "et0_qc",
        "et0_v31_yr_raw_valid_pixel_count",
        "et0_v31_yr_raw_area_weighted_mean_raw",
        "et0_v31_yr_sd_raw_valid_pixel_count",
        "et0_v31_yr_sd_raw_area_weighted_mean_raw",
    }
    _require(et0_fields.issubset(data.columns), "Author-approved ET0 fields were removed")
    _require(int(q_pass(data["et0_qc"]).sum()) == 3418, "ET0 retained-pass count changed")
    _require(not contract["source_field"].isin(et0_fields).any(), "ET0 entered the 21-estimand contract")
    _require(not facts["estimand"].isin(et0_fields).any(), "ET0 entered the 29 numerical facts")
    return {
        "status": "PASS",
        "shape": list(data.shape),
        "oasisid_unique": True,
        "class_counts": class_counts,
        "continent_counts": continent_counts,
        "spatial_blocks": blocks,
        "utci_complete": utci_pass,
        "utci_excluded": len(data) - utci_pass,
        "nex_methods": nex_methods,
        "futurepop": futurepop,
        "ai_scaled_valid": int(ai_values.notna().sum()),
        "et0": {
            "status": "approved_and_retained_raw_unit_caution",
            "valid": int(q_pass(data["et0_qc"]).sum()),
            "missing": int((~q_pass(data["et0_qc"])).sum()),
            "in_primary_estimands": False,
            "in_numeric_facts": False,
        },
    }


def _metadata_qa(repo_root: Path, data: pd.DataFrame) -> dict[str, object]:
    dictionary = pd.read_csv(repo_root / "data/data_dictionary.csv")
    _require(len(dictionary) == 44, "Data dictionary does not contain 44 fields")
    _require(set(dictionary["field"]) == set(data.columns), "Data dictionary field coverage changed")
    _require(
        dictionary["public_file_licence"].eq("CC-BY-4.0").all(),
        "Derived-data licence coverage is incomplete",
    )
    et0_rows = dictionary[dictionary["field"].str.startswith("et0_")]
    _require(len(et0_rows) == 5, "ET0 dictionary coverage changed")
    _require(
        et0_rows.loc[et0_rows["field"].eq("et0_qc"), "role"].eq(
            "retained_raw_unit_caution_qc"
        ).all()
        and et0_rows.loc[~et0_rows["field"].eq("et0_qc"), "role"].eq(
            "retained_raw_unit_caution"
        ).all(),
        "ET0 raw/unit-caution roles changed",
    )
    manifest = pd.read_csv(repo_root / "data/source_product_manifest.csv", dtype=str)
    redistributed = manifest["raw_files_redistributed"].str.lower().fillna("")
    _require(redistributed.eq("false").all(), "A source product is marked as raw-data redistributed")
    return {
        "data_dictionary_rows": len(dictionary),
        "data_dictionary_complete": True,
        "source_products": len(manifest),
        "raw_third_party_files_redistributed": False,
        "mixed_licence_policy": "code=MIT; derived_data_and_documentation=CC-BY-4.0",
    }


def _resume_result(output_dir: Path, fingerprint: dict[str, str]) -> dict[str, Any] | None:
    state_path = output_dir / "reproduction_state.json"
    qa_path = output_dir / "reproduction_QA.json"
    if not state_path.exists() or not qa_path.exists():
        return None
    try:
        state = json.loads(state_path.read_text(encoding="utf-8"))
        qa = json.loads(qa_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if state.get("status") == "PASS" and state.get("resume_fingerprint_sha256") == fingerprint:
        qa["resumed"] = True
        return qa
    return None


def validate_repository(
    repo_root: Path, output_dir: Path, *, resume: bool = False
) -> dict[str, Any]:
    repo_root = repo_root.resolve()
    output_dir = output_dir.resolve()
    files = {name: repo_root / relative for name, relative in REQUIRED_FILES.items()}
    missing = [str(path.relative_to(repo_root)) for path in files.values() if not path.is_file()]
    _require(not missing, f"Required repository files are missing: {missing}")
    input_hashes = {name: sha256(path) for name, path in files.items()}
    hash_mismatches = {
        name: {"expected": expected, "actual": input_hashes.get(name)}
        for name, expected in AUTHORITATIVE_RELEASE_SHA256.items()
        if input_hashes.get(name) != expected
    }
    _require(not hash_mismatches, f"Authoritative release hash mismatch: {hash_mismatches}")
    implementation_paths = [
        repo_root / "environment.yml",
        repo_root / "scripts/reproduce.py",
        *sorted((repo_root / "src").glob("*.py")),
        *sorted((repo_root / "tests").glob("*.py")),
    ]
    implementation_missing = [str(path.relative_to(repo_root)) for path in implementation_paths if not path.is_file()]
    _require(not implementation_missing, f"Implementation files are missing: {implementation_missing}")
    resume_fingerprint = dict(input_hashes)
    resume_fingerprint.update(
        {
            f"implementation:{path.relative_to(repo_root).as_posix()}": sha256(path)
            for path in implementation_paths
        }
    )
    if resume:
        resumed = _resume_result(output_dir, resume_fingerprint)
        if resumed is not None:
            return resumed

    config = yaml.safe_load(files["analysis_config"].read_text(encoding="utf-8"))
    contract = pd.read_csv(files["estimand_contract"])
    data = pd.read_csv(files["minimum_input"], low_memory=False)
    primary = pd.read_csv(files["primary_estimands"])
    facts = pd.read_csv(files["numeric_facts"])
    primary.attrs["source_path"] = str(files["primary_estimands"])
    facts.attrs["source_path"] = str(files["numeric_facts"])

    output_dir.mkdir(parents=True, exist_ok=True)
    state_path = output_dir / "reproduction_state.json"
    log_path = output_dir / "reproduction.log"
    work_names = (
        "point_estimands",
        "point_integrity",
        "locked_outputs",
        "minimum_input",
        "metadata",
        "source_data_workbook",
        "configuration",
    )
    previous_state = _load_interrupted_state(state_path, resume_fingerprint, resume)
    continuing = previous_state is not None
    if previous_state is None:
        state: dict[str, Any] = {
            "status": "running",
            "input_sha256": input_hashes,
            "resume_fingerprint_sha256": resume_fingerprint,
            "work_packages": {
                name: {
                    "status": "pending",
                    **(
                        {
                            "items": {
                                str(row["fact_id"]): {
                                    "status": "pending",
                                    "estimand": str(row["estimand"]),
                                }
                                for row in contract.to_dict("records")
                            }
                        }
                        if name == "point_estimands"
                        else {}
                    ),
                }
                for name in work_names
            },
            "outputs": {},
        }
        atomic_bytes(log_path, b"")
    else:
        state = previous_state
        state["status"] = "running"
        for name in work_names:
            state.setdefault("work_packages", {}).setdefault(name, {"status": "pending"})
        state["work_packages"]["point_estimands"].setdefault("items", {})
    atomic_json(state_path, state)
    _append_log(log_path, f"validation running resume={continuing}")

    def run_step(name: str, function: Any) -> dict[str, object]:
        work = state["work_packages"][name]
        result_path = output_dir / "step_results" / f"{name}.json"
        if (
            resume
            and work.get("status") == "success"
            and result_path.is_file()
            and work.get("result_sha256") == sha256(result_path)
        ):
            try:
                result = json.loads(result_path.read_text(encoding="utf-8"))
                _append_log(log_path, f"step resumed {name}")
                return result
            except (OSError, json.JSONDecodeError):
                pass
        work["status"] = "running"
        state["status"] = "running"
        atomic_json(state_path, state)
        _append_log(log_path, f"step running {name}")
        try:
            result = function()
            atomic_json(result_path, result)
            work.update(
                {
                    "status": "success",
                    "result_sha256": sha256(result_path),
                    "error": None,
                }
            )
            atomic_json(state_path, state)
            _append_log(log_path, f"step success {name}")
            return result
        except Exception as exc:
            work.update({"status": "failed", "error": f"{type(exc).__name__}: {exc}"})
            state["status"] = "failed"
            atomic_json(state_path, state)
            _append_log(log_path, f"step failed {name} {type(exc).__name__}: {exc}")
            raise

    regenerated, point_failures, resumed_estimators = _resumable_point_estimands(
        data,
        contract,
        output_dir,
        state,
        log_path,
        resume=resume,
    )
    point_qa = run_step(
        "point_integrity",
        lambda: _point_estimate_qa(
            regenerated, point_failures, primary, resumed_estimators
        )[1],
    )
    locked_qa = run_step(
        "locked_outputs", lambda: _locked_output_qa(contract, primary, facts)
    )
    data_qa = run_step("minimum_input", lambda: _data_qa(data, contract, facts))
    metadata_qa = run_step("metadata", lambda: _metadata_qa(repo_root, data))
    workbook_qa = run_step(
        "source_data_workbook",
        lambda: _workbook_qa(files["source_data"], primary, facts),
    )

    def configuration_qa() -> dict[str, object]:
        _require(config["expected"]["estimands"] == 21, "Configuration estimand count changed")
        _require(config["expected"]["numeric_facts"] == 29, "Configuration fact count changed")
        _require(config["aridity_index"]["scale_factor"] == 0.0001, "AI scale factor changed")
        return {
            "status": "PASS",
            "estimands": 21,
            "numeric_facts": 29,
            "ai_scale_factor": 0.0001,
        }

    configuration = run_step("configuration", configuration_qa)
    atomic_csv(output_dir / "regenerated_point_estimands_21.csv", regenerated)
    qa: dict[str, Any] = {
        "status": "PASS",
        "resumed": continuing,
        "repository": "global-oasis-exposure",
        "input_sha256": input_hashes,
        "authoritative_release_hashes_verified": True,
        "resume_fingerprint_sha256": resume_fingerprint,
        "point_estimates": point_qa,
        "locked_outputs": locked_qa,
        "minimum_input": data_qa,
        "metadata": metadata_qa,
        "source_data_workbook": workbook_qa,
        "configuration": configuration,
        "summary": {
            "estimands": 21,
            "robust": 16,
            "sensitive": 2,
            "not_supported": 3,
            "numeric_facts": 29,
            "bootstrap_recomputed": False,
        },
    }
    atomic_json(output_dir / "reproduction_QA.json", qa)
    _append_log(log_path, "validation PASS")
    state["status"] = "PASS"
    state["outputs"] = {
        "regenerated_point_estimands_21.csv": sha256(
            output_dir / "regenerated_point_estimands_21.csv"
        ),
        "reproduction_QA.json": sha256(output_dir / "reproduction_QA.json"),
        "reproduction.log": sha256(log_path),
    }
    atomic_json(output_dir / "reproduction_state.json", state)
    return qa
