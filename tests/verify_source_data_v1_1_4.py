"""Structural checks for the public v1.1.4 Source Data workbook.

This test checks only the shipped workbook's public shape and declared
weighting/threshold metadata. It does not rebuild figures or redistribute any
provider-controlled source product.

Run with:  python -m unittest tests.verify_source_data_v1_1_4 -v
"""

from __future__ import annotations

import unittest
import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data" / "Source_Data_CEE_v1_1_4.xlsx"

EXPECTED_SHEETS = (
    "README",
    "Data_Dictionary",
    "Figure2_Polygon",
    "Figure2_Summary",
    "Figure3_Polygon",
    "Figure3_ClassSummary",
    "Figure3_AreaBins",
    "Figure4_Polygon",
    "Figure4_Summary",
    "Figure5_Polygon",
    "Figure5_ClassSummary",
    "Figure6_Polygon",
    "Figure6_A_Points",
    "Figure6_A_ClassMeans",
    "Figure6_B_PrimaryThreshold",
    "Figure6_C_MultiThreshold",
    "Figure7_Polygon",
    "Figure7_ClassSummary",
    "Figure7_RegionSummary",
    "Figure8_Polygon",
    "Figure8_Summary",
    "Primary_Estimands_31",
    "NEX_Model_IQR_24",
    "FuturePop_Sensitivity_20",
    "FuturePop_CoverageSummary",
    "SuppTableS1_GDP",
    "SuppTableS5_NEX",
    "SuppTableS7_Products",
    "SuppTableS13_ProductBoundary",
    "SuppFigS1_Polygon",
    "SuppFigS1_ClassSummary",
)


class SourceDataV114Tests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.wb = load_workbook(WORKBOOK, read_only=True, data_only=True)

    @classmethod
    def tearDownClass(cls) -> None:
        cls.wb.close()

    def test_sheet_order_and_polygon_rows(self) -> None:
        self.assertEqual(tuple(self.wb.sheetnames), EXPECTED_SHEETS)
        for name in EXPECTED_SHEETS:
            self.assertEqual(self.wb[name].sheet_state, "visible", name)
        for name in EXPECTED_SHEETS:
            if name.endswith("_Polygon"):
                self.assertEqual(self.wb[name].max_row, 3438, name)

    def test_counts_and_nex_model_count(self) -> None:
        self.assertEqual(self.wb["Primary_Estimands_31"].max_row, 32)
        nex = self.wb["NEX_Model_IQR_24"]
        header = next(nex.iter_rows(min_row=1, max_row=1, values_only=True))
        model_idx = header.index("model_count")
        counts = {row[model_idx] for row in nex.iter_rows(min_row=2, values_only=True)}
        self.assertEqual(counts, {30})

    def test_futurepop_coverage_accounting(self) -> None:
        ws = self.wb["FuturePop_CoverageSummary"]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        frac_idx = header.index("polygon_fractional_oases")
        nearest_idx = header.index("nearest_valid_grid_oases")
        class_idx = header.index("background_class")
        unique_by_class = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            unique_by_class[row[class_idx]] = (row[frac_idx], row[nearest_idx])
        self.assertEqual(sum(item[0] for item in unique_by_class.values()), 3181)
        self.assertEqual(sum(item[1] for item in unique_by_class.values()), 256)

    def test_figure5_weighting_and_figure6_threshold_metadata(self) -> None:
        fig5 = self.wb["Figure5_ClassSummary"]
        header = next(fig5.iter_rows(min_row=1, max_row=1, values_only=True))
        weighting_idx = header.index("weighting_scheme")
        weighting = [row[weighting_idx] for row in fig5.iter_rows(min_row=2, values_only=True)]
        self.assertTrue(all("oasis-area-weighted" in str(value) for value in weighting))
        self.assertTrue(all("water-pixel-area-weighted" in str(value) for value in weighting))

        fig6 = self.wb["Figure6_B_PrimaryThreshold"]
        header = next(fig6.iter_rows(min_row=1, max_row=1, values_only=True))
        metric_idx = header.index("metric")
        metrics = {row[metric_idx] for row in fig6.iter_rows(min_row=2, values_only=True)}
        self.assertIn("heat >=32C", metrics)
        self.assertIn("cold <=-13C", metrics)

    def test_public_workbook_contains_no_hidden_formula_or_comment_state(self) -> None:
        for ws in self.wb.worksheets:
            self.assertEqual(ws.sheet_state, "visible", ws.title)
            for row in ws.iter_rows():
                for cell in row:
                    self.assertIsNone(
                        getattr(cell, "comment", None),
                        f"comment in {ws.title}!{cell.coordinate}",
                    )
                    self.assertIsNone(
                        getattr(cell, "hyperlink", None),
                        f"hyperlink in {ws.title}!{cell.coordinate}",
                    )
                    self.assertFalse(
                        isinstance(cell.value, str) and cell.value.startswith("="),
                        f"formula in {ws.title}!{cell.coordinate}",
                    )

    def test_public_dictionary_covers_analysis_columns(self) -> None:
        with (ROOT / "data" / "data_dictionary.csv").open(
            encoding="utf-8-sig", newline=""
        ) as handle:
            fields = {row["field"] for row in csv.DictReader(handle)}
        self.assertEqual(len(fields), 219)
        self.assertTrue(
            {
                "baseline_id",
                "block_250km",
                "block_500km",
                "block_1000km",
                "utci_denominator_method",
                "utci_invalid_years",
            }.issubset(fields)
        )


if __name__ == "__main__":
    unittest.main()
